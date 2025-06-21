import pandas as pd
import numpy as np
from tkinter import Tk
from tkinter.filedialog import askopenfilename
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00, FORMAT_NUMBER_COMMA_SEPARATED1

# Cargar archivo de pronósticos
Tk().withdraw()
ruta_forecast = askopenfilename(title="Selecciona el archivo de pronóstico", filetypes=[("Excel files", "*.xlsx")])
forecast_df = pd.read_excel(ruta_forecast)

# Cargar archivo de datos logísticos
ruta_datos = askopenfilename(title="Selecciona el archivo de datos maestros", filetypes=[("Excel files", "*.xlsx")])
datos_df = pd.read_excel(ruta_datos)

# Preprocesamiento
hoy = datetime.now()
sku_sitstand = [
    "Escritorio Sit&Stand"
]

capacidad_silla = 50 * 12
capacidad_sitstand = 5 * 25
capacidad_total = capacidad_silla + capacidad_sitstand
container_capacidad = 600

resultados = []
for _, row in datos_df.iterrows():
    sku = row["SKU"]
    stock_actual = row["Stock actual"]
    lead_time = row["Tiempo de espera"]
    stock_transito = row["Stock en transito"]
    fecha_llegada = row["Fecha de llegada de stock en transito"]
    rop = row["ROP"]
    ss = row["SS"]
    beneficio_unit = row["Beneficio x unidad"]

    pronostico = forecast_df[forecast_df["SKU"] == sku]["Pronóstico Próximos 6"].values
    if len(pronostico) == 0:
        continue
    pronostico = eval(pronostico[0])

    dias = lead_time
    demanda_durante_lead = 0
    dias_restantes = dias
    i = 0
    while dias_restantes > 0 and i < len(pronostico):
        dias_periodo = 30
        if dias_restantes >= dias_periodo:
            demanda_durante_lead += pronostico[i]
        else:
            demanda_durante_lead += (dias_restantes / dias_periodo) * pronostico[i]
        dias_restantes -= dias_periodo
        i += 1
    demanda_durante_lead = round(demanda_durante_lead)

    if pd.notna(fecha_llegada):
        dias_a_llegada = (pd.to_datetime(fecha_llegada) - hoy).days
        stock_en_camino = stock_transito if dias_a_llegada <= lead_time else 0
    else:
        stock_en_camino = 0

    stock_proyectado = max(0, stock_actual - demanda_durante_lead + stock_en_camino)

    dias_restantes = lead_time
    i = 0
    while dias_restantes > 0 and i < len(pronostico):
        dias_restantes -= 30
        i += 1
    periodo_usado = pronostico[i:i+3]
    stock_objetivo = max(0, sum(periodo_usado) + ss)

    # --------- Validación añadida ---------
    if stock_actual < rop or stock_proyectado < ss:
        minimo_compra = max(0, stock_objetivo - stock_proyectado)
    else:
        minimo_compra = 0
    # --------------------------------------

    beneficio_total = beneficio_unit * minimo_compra
    tipo = "SitStand" if sku in sku_sitstand else "Silla"
    rack_usado = round(minimo_compra / (25 if tipo == "SitStand" else 12), 2)

    resultados.append({
        "SKU": sku,
        "Cantidad a comprar": minimo_compra,
        "Beneficio x unidad": beneficio_unit,
        "Beneficio total": beneficio_total,
        "Tipo": tipo,
        "Racks usados": rack_usado
    })

df = pd.DataFrame(resultados).sort_values(by="Beneficio total", ascending=False)

acum_container = 0
acum_silla = 0
acum_sitstand = 0
acum_total_rack = 0
container_data = []

for _, row in df.iterrows():
    cant = row["Cantidad a comprar"]
    tipo = row["Tipo"]
    rack = row["Racks usados"]

    pct_container = (cant / container_capacidad)
    pct_container_acum = acum_container + pct_container

    if tipo == "Silla":
        pct_almacen_tipo = (cant / capacidad_silla)
        pct_almacen_tipo_acum = acum_silla + pct_almacen_tipo
        acum_silla += pct_almacen_tipo
    else:
        pct_almacen_tipo = (cant / capacidad_sitstand)
        pct_almacen_tipo_acum = acum_sitstand + pct_almacen_tipo
        acum_sitstand += pct_almacen_tipo

    pct_total_rack = ((acum_total_rack + rack) / (capacidad_silla / 12 + capacidad_sitstand / 25))
    acum_total_rack += rack
    acum_container += pct_container

    container_data.append({
        **row,
        "% almacén silla": pct_almacen_tipo if tipo == "Silla" else 0,
        "% almacén silla acum": pct_almacen_tipo_acum if tipo == "Silla" else acum_silla,
        "% almacén sitstand": pct_almacen_tipo if tipo == "SitStand" else 0,
        "% almacén sitstand acum": pct_almacen_tipo_acum if tipo == "SitStand" else acum_sitstand,
        "% almacén total acum": pct_total_rack,
        "% container": pct_container,
        "% container acum": pct_container_acum
    })

df_export = pd.DataFrame(container_data)
archivo_salida = "visso-cantidad-comprar.xlsx"
df_export.to_excel(archivo_salida, index=False)

# Formateo con openpyxl
wb = load_workbook(archivo_salida)
ws = wb.active

cabeceras = list(df_export.columns)
azul = PatternFill(start_color="DEEDF2", end_color="DEEDF2", fill_type="solid")
rojo = PatternFill(start_color="EDB1A6", end_color="EDB1A6", fill_type="solid")
font_bold = Font(bold=True)
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

ws.column_dimensions["A"].width = max(15, max(len(str(cell.value)) for cell in ws["A"]))
align_left = Alignment(horizontal="left", vertical="distributed", wrap_text=True)
for cell in ws["A"]:
    cell.alignment = align_left
for col in range(2, len(cabeceras)+1):
    letra = get_column_letter(col)
    ws.column_dimensions[letra].width = 10
    for cell in ws[letra]:
        cell.alignment = align_center

for cell in ws[1]:
    cell.font = font_bold
    cell.alignment = align_center

for row_idx in range(2, ws.max_row + 1):
    container_acum = ws[f"M{row_idx}"].value
    total_rack_acum = ws[f"L{row_idx}"].value
    fill = azul if container_acum <= 1 and total_rack_acum <= 0.9 else rojo
    for col in range(1, len(cabeceras)+1):
        cell = ws.cell(row=row_idx, column=col)
        cell.alignment = align_center
        if "%" in cabeceras[col-1]:
            cell.number_format = FORMAT_PERCENTAGE_00
        elif "Beneficio" in cabeceras[col-1]:
            cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
        cell.fill = fill

wb.save(archivo_salida)
print(f"Archivo exportado con formato: {archivo_salida}")
